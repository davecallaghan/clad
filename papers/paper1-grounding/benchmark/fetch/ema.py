"""EMA: authorised medicines, indications, and authorisation lifecycle dates.

Supplies the jurisdiction family, and independently supplies EU-side supersession
via withdrawal, revocation, and suspension dates.

Two facts about coverage shape what can be built from it. EMA's register is of
*centrally* authorised medicines, which skews modern -- biologics, oncology, rare
disease -- while older generics were authorised nationally and do not appear. So
the FDA-EMA substance overlap is real but narrower than the size of either
register suggests, and it concentrates in exactly the high-consequence products
where an applicability error costs most.

REUSE: the EMA download page states no licence or reuse terms. Verify before
publishing derived data.
"""
import io, re
import openpyxl
from common import get

XLSX = ("https://www.ema.europa.eu/en/documents/report/"
        "medicines-output-medicines-report_en.xlsx")

FIELDS = {
    "medicine": "Name of medicine", "status": "Medicine status",
    "inn": "International non-proprietary name (INN) / common name",
    "substance": "Active substance", "indication": "Therapeutic indication",
    "atc": "ATC code (human)", "revision": "Revision number",
    "auth_date": "Marketing authorisation date",
    "withdrawn": "Withdrawal / expiry / revocation / lapse of marketing authorisation date",
    "suspended": "Suspension of marketing authorisation date",
    "first_published": "First published date", "last_updated": "Last updated date",
    "url": "Medicine URL",
}


def load():
    wb = openpyxl.load_workbook(io.BytesIO(get(XLSX, binary=True)),
                                read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    hdr = None
    for r in rows:
        if r and any(str(x) == "Name of medicine" for x in r if x):
            hdr = [str(x).replace("\n", " ").strip() if x else "" for x in r]
            break
    if hdr is None:
        return []
    col = {}
    for key, label in FIELDS.items():
        hit = next((i for i, h in enumerate(hdr)
                    if h == label or h.startswith(label[:38])), None)
        if hit is not None:
            col[key] = hit
    out = []
    for r in rows:
        if not r or col.get("medicine") is None or not r[col["medicine"]]:
            continue
        rec = {k: (str(r[i]).strip() if i < len(r) and r[i] is not None else None)
               for k, i in col.items()}
        out.append(rec)
    return out


def norm(s):
    """Normalise a substance name for cross-register matching."""
    s = (s or "").lower().strip()
    s = re.sub(r"\s*\(.*?\)\s*", " ", s)            # drop parentheticals
    s = re.sub(r"[^a-z0-9 \-]", " ", s)
    return " ".join(s.split())


def substance_index(rows):
    """normalised substance -> [record]. Multi-substance products index under each."""
    idx = {}
    for r in rows:
        raw = r.get("substance") or r.get("inn") or ""
        # Combination products separate on newlines or commas; single INNs may
        # legitimately contain a hyphen or space, so do not split on those.
        for part in re.split(r"[\n;]|,\s", raw):
            k = norm(part)
            if len(k) > 3:
                idx.setdefault(k, []).append(r)
    return idx
